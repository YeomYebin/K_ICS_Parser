"""추출 결과를 예시 표 형식의 .xlsx 로 생성."""
import io

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


HEADER_FILL = PatternFill("solid", fgColor="D9E1F2")
WARN_FILL = PatternFill("solid", fgColor="FFC7CE")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT_WRAP = Alignment(horizontal="left", vertical="top", wrap_text=True)


def build_excel(rows, quarter_label="'26.03월"):
    wb = Workbook()
    ws = wb.active
    ws.title = "K-ICS"

    headers = [quarter_label, "K-ICS 비율", "가용자본", "요구자본",
               "기본자본", "보완자본", "주요 변동요인"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = BORDER

    for r in rows:
        ratio_shown = r.get("kics_ratio_display")
        if ratio_shown is None:
            ratio_shown = r.get("kics_ratio")
        ws.append([
            r.get("display_name") or r.get("company"),
            ratio_shown,  # 표시값 = 가용÷요구 계산값 (없으면 문서값)
            r.get("available_capital"),
            r.get("required_capital"),
            r.get("tier1_capital"),
            r.get("tier2_capital"),
            r.get("change_factors") or "",
        ])

    max_row = ws.max_row
    for row_idx in range(2, max_row + 1):
        rec = rows[row_idx - 2]
        for col_idx in range(1, 8):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = BORDER
            if col_idx == 1:
                cell.alignment = CENTER
            elif col_idx == 2:
                cell.number_format = "0.0"
                cell.alignment = CENTER
            elif 3 <= col_idx <= 6:
                cell.number_format = "#,##0"
                cell.alignment = Alignment(horizontal="right", vertical="center")
            else:
                cell.alignment = LEFT_WRAP
        # 계산값(표시값)이 문서 표기값과 다르면 K-ICS 비율 셀을 빨간색으로 강조
        if rec.get("ratio_warning"):
            wc = ws.cell(row=row_idx, column=2)
            wc.fill = WARN_FILL
            wc.font = Font(bold=True, color="C00000")  # 빨간색 글씨
            calc = rec.get("kics_ratio_calc")
            doc = rec.get("kics_ratio")
            from openpyxl.comments import Comment
            wc.comment = Comment(
                f"계산값(가용/요구*100) = {calc}\n문서 표기값 = {doc}", "검산")

    widths = [12, 11, 12, 12, 12, 12, 60]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

    ws.freeze_panes = "A2"

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio
